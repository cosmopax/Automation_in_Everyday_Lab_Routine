import os
import re
import argparse
import logging
import pubchempy as pcp
from rdkit import Chem
from rdkit.Chem import AllChem, Draw
from googlesearch import search
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image

# www.GitHub.com/cosmopax

# Configure logging
def setup_logging():
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s [%(levelname)s] %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )

# Utilities
def sanitize_filename(name):
    return re.sub(r'[^\w\-_. ]', '_', name)

# Determine base directory dynamically
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Chemical data functions
def google_for_iupac_or_cas(colloquial_name):
    logging.info(f"Searching for '{colloquial_name}' on Google...")
    query = f"{colloquial_name} IUPAC or CAS number"
    for result in search(query, num_results=5):
        logging.debug(f"Checking result: {result}")
        cas_match = re.search(r'\b\d{2,7}-\d{2}-\d\b', result)
        if cas_match:
            logging.info(f"Found CAS: {cas_match.group()}")
            return cas_match.group()
        if "pubchem.ncbi.nlm.nih.gov/compound/" in result.lower():
            cid = result.rstrip('/').split('/')[-1]
            logging.info(f"Found PubChem CID: {cid}")
            return cid
    logging.warning("No CAS or CID found via Google.")
    return None

# Fetch chemical details from PubChem
def fetch_details_from_pubchem(query, query_type='name'):
    logging.info(f"Fetching PubChem data: query={query}, type={query_type}")
    try:
        if query_type == 'name':
            comps = pcp.get_compounds(query, 'name')
        elif query_type == 'cid':
            comps = pcp.get_compounds(query, 'cid')
        else:
            raise ValueError("query_type must be 'name' or 'cid'")
        if not comps:
            raise LookupError(f"No compound for {query}")
        c = comps[0]
        smiles = c.isomeric_smiles
        iupac = c.iupac_name or ''
        try:
            mw = float(c.molecular_weight)
        except Exception:
            mw = None
        formula = c.molecular_formula or ''
        synonyms = c.synonyms[:5] if c.synonyms else []
        # Manual override for known compounds
        lowq = query.lower()
        if lowq in ['2c-b','2-(4-bromo-2,5-dimethoxyphenyl)ethanamine']:
            smiles = 'Cc1c(OC)cc(Br)cc1OCCCN'
        logging.info(f"SMILES: {smiles}")
        return smiles, iupac, mw, formula, synonyms
    except Exception as e:
        logging.error(f"PubChem fetch failed: {e}")
        return None, None, None, None, None

# Generate structure files
def generate_structures(smiles, pdb_dir, mol_dir, iupac, colloq):
    base = f"{sanitize_filename(iupac)}_{sanitize_filename(colloq)}"
    pdb_dir = os.path.expanduser(pdb_dir)
    mol_dir = os.path.expanduser(mol_dir)
    os.makedirs(pdb_dir, exist_ok=True)
    os.makedirs(mol_dir, exist_ok=True)
    pdb_path = os.path.join(pdb_dir, base + '.pdb')
    mol_path = os.path.join(mol_dir, base + '.mol')
    try:
        mol = Chem.MolFromSmiles(smiles)
        mol_h = Chem.AddHs(mol)
        AllChem.EmbedMolecule(mol_h, AllChem.ETKDG())
        AllChem.UFFOptimizeMolecule(mol_h)
        mol_clean = Chem.RemoveHs(mol_h)
        with open(pdb_path,'w') as f:
            f.write(Chem.MolToPDBBlock(mol_clean))
        Chem.MolToMolFile(mol_clean, mol_path)
        logging.info(f"Saved PDB: {pdb_path}")
        logging.info(f"Saved MOL: {mol_path}")
        return pdb_path, mol_path
    except Exception as e:
        logging.error(f"Structure generation failed: {e}")
        return None, None

# Excel logging
def log_to_excel(xl_path, img_path, colloq, iupac, mw, formula, synonyms):
    xl_path = os.path.expanduser(xl_path)
    os.makedirs(os.path.dirname(xl_path), exist_ok=True)
    if os.path.exists(xl_path):
        wb = load_workbook(xl_path)
    else:
        wb = Workbook()
    ws = wb.active
    if ws.max_row==1 and ws.max_column==1:
        ws.append(["Laufnummer","Structure","Colloquial Name","IUPAC Name","Molar Mass (g/mol)","Formula","Melting Pt (°C)","NMR Link","MS Link","Synonyms"])
    num = ws.max_row
    for row in ws.iter_rows(min_row=2,values_only=True):
        if row[3]==iupac:
            logging.info("Excel entry exists, skipping")
            return
    ws.append([num, None, colloq, iupac, mw, formula, "","","", "; ".join(synonyms)])
    if img_path:
        img = Image(img_path)
        img.anchor = f"B{ws.max_row}"
        ws.add_image(img)
    wb.save(xl_path)
    logging.info(f"Logged to Excel: {xl_path}")

# 2D image generation
def draw_2d(smiles, img_dir, iupac, colloq):
    img_dir = os.path.expanduser(img_dir)
    base = f"{sanitize_filename(iupac)}_{sanitize_filename(colloq)}.png"
    os.makedirs(img_dir, exist_ok=True)
    path = os.path.join(img_dir, base)
    mol = Chem.MolFromSmiles(smiles)
    AllChem.Compute2DCoords(mol)
    Draw.MolToFile(mol, path, kekulize=True)
    return path

# Command-line interface
def main():
    setup_logging()
    parser = argparse.ArgumentParser(description="Chemical data fetcher")
    parser.add_argument("-e","--excel", default=os.path.join(BASE_DIR, "neurochem.xlsx"), help="Excel file path")
    parser.add_argument("-p","--pdb_dir", default=os.path.join(BASE_DIR, "ligands"), help="PDB output dir")
    parser.add_argument("-m","--mol_dir", default=os.path.join(BASE_DIR, "chemdraw"), help="MOL output dir")
    parser.add_argument("-i","--img_dir", default=os.path.join(BASE_DIR, "images"), help="2D image dir")
    args = parser.parse_args()
    excel = args.excel
    pdb_dir = args.pdb_dir
    mol_dir = args.mol_dir
    img_dir = args.img_dir

    logging.info("Enter Ctrl+C to exit or type 'exit'")
    while True:
        try:
            colloq = input("Compound name or CAS: ").strip()
            if colloq.lower()=='exit': break
            query = google_for_iupac_or_cas(colloq) or colloq
            smiles,iupac,mw,formula,synonyms = fetch_details_from_pubchem(query,'name')
            if smiles:
                pdb_path,mol_path = generate_structures(smiles,pdb_dir,mol_dir,iupac,colloq)
                img_path = draw_2d(smiles,img_dir,iupac,colloq)
                log_to_excel(excel,img_path,colloq,iupac,mw,formula,synonyms)
        except KeyboardInterrupt:
            break
    logging.info("Goodbye")

if __name__=="__main__":
    main()
